import sys
import os
from PyQt5 import QtWidgets, QtGui
from PyQt5.QtWidgets import QFileDialog, QMessageBox, QRadioButton, QButtonGroup
import pandas as pd
import numpy as np
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
import matplotlib.dates as mdates
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import column_index_from_string
import matplotlib.pyplot as plt
import matplotlib
matplotlib.use('Qt5Agg')


def process_interval_df(df):
    # Spojení sloupců Datum a Čas do jednoho datetime sloupce
    df['Datetime'] = pd.to_datetime(df['Date'] + ' ' + df['Time'], dayfirst=True, format='%d.%m.%Y %H:%M:%S.%f')

    # Posunout hodnoty deformace, aby první hodnota byla nula
    initial_deform = df['Deformace'].iloc[0]
    df['Deformace'] = df['Deformace'] - initial_deform

    # Inverze znamének pro sílu
    df['Síla'] = -df['Síla']

    # Najít index maximální síly (po invertu to bude minimum původní síly) a odříznout předchozí řádky
    idx_max_force = df['Síla'].idxmax()
    df_trim = df.loc[idx_max_force:].reset_index(drop=True)

    # Uložit hodnotu deformace v okamžiku maxima síly
    deform_at_max = df_trim['Deformace'].iloc[0]
    # Nastavit počet bodů, po kterých kontrolujeme návrat deformace
    window = 10
    idx_end = None
    # Projít body od indexu `window` dále a najít první, kde se deformace vrátí k hodnotě nebo je menší
    for i in range(window, len(df_trim)):
        if df_trim['Deformace'].iloc[i] <= deform_at_max:
            idx_end = i
            break
    if idx_end is not None:
        df_trim = df_trim.loc[:idx_end].reset_index(drop=True)

    # Vypočítat uplynulý čas od začátku zkrácených dat (v sekundách)
    df_trim['ElapsedTime'] = (df_trim['Datetime'] - df_trim.loc[0, 'Datetime']).dt.total_seconds()

    # Připravit pole pro fitování
    time_vals = df_trim['ElapsedTime'].values
    force_vals = df_trim['Síla'].values
    deform_vals = df_trim['Deformace'].values

    # Stanovit maximální stupeň polynomu (nejvýše 3, méně pokud málo bodů)
    unique_times = np.unique(time_vals)
    max_deg = min(3, len(unique_times) - 1)

    # Inicializace proměnných
    coeffs_force = None
    coeffs_deform = None
    df_trim['Síla_fit'] = np.nan
    df_trim['Deformace_fit'] = np.nan
    used_deg = 0

    if max_deg >= 1:
        for deg in range(max_deg, 0, -1):
            try:
                coeffs_force = np.polyfit(time_vals, force_vals, deg=deg)
                coeffs_deform = np.polyfit(time_vals, deform_vals, deg=deg)
                df_trim['Síla_fit'] = np.polyval(coeffs_force, time_vals)
                df_trim['Deformace_fit'] = np.polyval(coeffs_deform, time_vals)
                used_deg = deg
                break
            except np.linalg.LinAlgError:
                continue

    # Výpočet inflexního bodu (pouze pro kubický polynom)
    infl_time = None
    infl_force = None
    infl_deform = None
    infl_force_fit = None
    infl_deform_fit = None
    if used_deg == 3 and coeffs_force is not None:
        a, b, c, d = coeffs_force
        # Druhá derivace f''(t) = 6a t + 2b => t = -b/(3a)
        infl_time = -b / (3 * a)
        # Najdeme index nejbližší hodnotě infl_time
        idx_nearest = np.argmin(np.abs(time_vals - infl_time))
        # Získat reálné hodnoty síly a deformace v tomto čase
        infl_force = df_trim.loc[idx_nearest, 'Síla']
        infl_deform = df_trim.loc[idx_nearest, 'Deformace']
        # Hodnoty z nafitované křivky
        infl_force_fit = np.polyval(coeffs_force, infl_time)
        infl_deform_fit = np.polyval(coeffs_deform, infl_time)

    return (df_trim.drop(columns=['Datetime']), used_deg, infl_time, infl_force,
            infl_deform, infl_force_fit, infl_deform_fit, coeffs_force, coeffs_deform)


# --- Matplotlib canvas ---
class MplCanvas(FigureCanvas):
    def __init__(self, parent=None, width=15, height=5, dpi=100):
        fig = Figure(figsize=(width, height), dpi=dpi)
        self.ax = fig.add_subplot(111)
        super().__init__(fig)

# --- Main application ---
class SCBApp(QtWidgets.QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle('SCB Test Excel Exporter')
        font = QtGui.QFont()
        font.setPointSize(10)
        self.setFont(font)
        self.setGeometry(100, 100, 1600, 900)
        self.df = pd.DataFrame()
        self.segments = []
        self.sample_params = None
        if hasattr(sys, '_MEIPASS'):
            base_path = sys._MEIPASS
        else:
            base_path = os.path.abspath(".")

        self.template_path = os.path.join(base_path, 'scb_sablona.xlsx')
        self.output_path = ''
        self._init_ui()

    def _init_ui(self):
        widget = QtWidgets.QWidget()
        self.setCentralWidget(widget)
        vbox = QtWidgets.QVBoxLayout(widget)
        # controls
        hbox = QtWidgets.QHBoxLayout()
        vbox.addLayout(hbox)
        btn_load = QtWidgets.QPushButton('Load Data File')
        btn_load.clicked.connect(self.load_file)
        self.chk_zoom = QtWidgets.QCheckBox('Zoom on drag')
        self.chk_zoom.setChecked(False)
        hbox.addWidget(btn_load)
        hbox.addWidget(self.chk_zoom)

        # Parameters
        btn_params = QtWidgets.QPushButton('Sample parameters')
        btn_params.clicked.connect(self.edit_parameters)
        hbox.addWidget(btn_params)

        # Radiobuttony pro volbu metody inflexních bodů
        self.radio_auto = QRadioButton("SW inflex point")
        self.radio_preserve = QRadioButton("Excel inflex point")
        self.radio_auto.setChecked(True)
        grp = QButtonGroup(self)
        grp.addButton(self.radio_auto)
        grp.addButton(self.radio_preserve)
        hbox.addWidget(self.radio_auto)
        hbox.addWidget(self.radio_preserve)

        # Threshold spinbox
        hbox.addWidget(QtWidgets.QLabel('Force threshold:'))
        self.thr_spin = QtWidgets.QDoubleSpinBox()
        self.thr_spin.setRange(0,0.2)
        self.thr_spin.setSingleStep(0.01)
        self.thr_spin.setValue(0.03)
        hbox.addWidget(self.thr_spin)

        # MA window spinbox
        hbox.addWidget(QtWidgets.QLabel('MA window:'))
        self.ma_spin = QtWidgets.QSpinBox()
        self.ma_spin.setRange(1,50)
        self.ma_spin.setValue(20)
        hbox.addWidget(self.ma_spin)

        # MA threshold spinbox
        hbox.addWidget(QtWidgets.QLabel('MA threshold:'))
        self.ma_thr_spin = QtWidgets.QDoubleSpinBox()
        self.ma_thr_spin.setRange(0,0.2)
        self.ma_thr_spin.setSingleStep(0.01)
        self.ma_thr_spin.setValue(0.05)
        hbox.addWidget(self.ma_thr_spin)

        # ← Sem pod přidání těchto řádků:
        self.thr_spin.valueChanged.connect(self.on_params_changed)
        self.ma_spin.valueChanged.connect(self.on_params_changed)
        self.ma_thr_spin.valueChanged.connect(self.on_params_changed)

        hbox.addWidget(QtWidgets.QLabel('Start DateTime:'))
        self.start_edit = QtWidgets.QLineEdit(); hbox.addWidget(self.start_edit)
        hbox.addWidget(QtWidgets.QLabel('End DateTime:'))
        self.end_edit = QtWidgets.QLineEdit(); hbox.addWidget(self.end_edit)
        btn_update = QtWidgets.QPushButton('Update View'); btn_update.clicked.connect(self.update_plot)
        btn_export = QtWidgets.QPushButton('Export to Excel'); btn_export.clicked.connect(self.export_selection)
        hbox.addWidget(btn_update)
        hbox.addWidget(btn_export)
        # plot
        self.canvas = MplCanvas(self, width=10, height=5, dpi=100)
        vbox.addWidget(self.canvas)
        self.canvas.mpl_connect('button_press_event', self.on_canvas_click)

    def load_file(self):
        path, _ = QFileDialog.getOpenFileName(self, 'Select data file', '', 'CSV/TXT (*.csv *.txt);;All (*)')
        if not path:
            return
        df = pd.read_csv(path, sep=';', skiprows=6, header=None, decimal=',')
        df.columns = ['datum','čas','ignore','síla','deformace']
        df.drop(columns=['ignore'], inplace=True)
        df['síla'] = pd.to_numeric(df['síla'], errors='coerce')
        df['deformace'] = pd.to_numeric(df['deformace'], errors='coerce')
        df['datetime'] = pd.to_datetime(df['datum'].astype(str)+' '+df['čas'].astype(str), format='%d.%m.%y %H:%M:%S.%f', errors='coerce')
        df.dropna(subset=['datetime','síla','deformace'], inplace=True)
        self.df = df.sort_values('datetime').reset_index(drop=True)
        fmt = lambda dt: dt.strftime('%Y-%m-%d %H:%M:%S.%f')[:-5]
        self.start_edit.setText(fmt(self.df.datetime.iloc[0]))
        self.end_edit.setText(fmt(self.df.datetime.iloc[-1]))
        self._detect_segments()
        self._draw_plot()

    def _detect_segments(self):
        thr = self.thr_spin.value()
        ma_win = self.ma_spin.value()
        ma_thr = self.ma_thr_spin.value()
        df = self.df
        if df.empty:
            self.segments = []
            return
        abs_f = df['síla'].abs()
        ma = abs_f.rolling(window=ma_win, min_periods=1, center=True).mean()
        mask = (abs_f > thr) & (ma > ma_thr)

        min_len = 20  # minimální počet bodů v intervalu
        segs, in_test = [], False
        for i, active in enumerate(mask):
            if active and not in_test:
                start = i
                in_test = True
            elif not active and in_test:
                end = i - 1
                if end - start + 1 >= min_len:
                    segs.append((start, end))
                in_test = False

        # pokud skončíme uvnitř intervalu, zkontrolujeme nakonec
        if in_test:
            end = len(df) - 1
            if end - start + 1 >= min_len:
                segs.append((start, end))

        self.segments = segs

    def on_params_changed(self):
        if self.df.empty:
            return
        self._detect_segments()
        self._draw_plot()

    def _draw_plot(self):
        ax = self.canvas.ax; ax.clear(); df = self.df
        try:
            s = pd.to_datetime(self.start_edit.text()); e = pd.to_datetime(self.end_edit.text())
            view = df[(df.datetime >= s) & (df.datetime <= e)]
        except:
            view = df
        ax.plot(view.datetime, view.síla, label='Síla', alpha=0.6)
        ax.plot(view.datetime, view.deformace, label='Deformace', alpha=0.6)
        for d in getattr(self, 'draggables', []):
            self.canvas.mpl_disconnect(d.cid_press); self.canvas.mpl_disconnect(d.cid_release); self.canvas.mpl_disconnect(d.cid_motion)
        self.draggables = []
        for idx, (st, ed) in enumerate(self.segments):
            dt_s, dt_e = df.datetime.iloc[st], df.datetime.iloc[ed]
            if view is not df and (dt_s < view.datetime.min() or dt_e > view.datetime.max()):
                continue
            xs, xe = mdates.date2num(dt_s), mdates.date2num(dt_e)
            ax.axvspan(dt_s, dt_e,
                       color=f'C{idx % 10}',  # nebo např. plt.cm.tab10(idx % 10)
                       alpha=0.2)
            ls = ax.axvline(xs, color='green', linestyle='--', picker=5)
            le = ax.axvline(xe, color='red', linestyle='--', picker=5)
            self.draggables.extend([DraggableVLine(ls, idx, True, self), DraggableVLine(le, idx, False, self)])
        ax.legend(); ax.set_xlabel('Time'); ax.set_ylabel('Value'); ax.grid(True)
        self.canvas.draw()

    def on_canvas_click(self, event):
        if event.button == 3 and event.xdata is not None:
            x0, x1 = self.canvas.ax.get_xlim(); thresh = (x1 - x0)*0.005
            for idx, (st, ed) in enumerate(self.segments):
                xs = mdates.date2num(self.df.datetime.iloc[st]); xe = mdates.date2num(self.df.datetime.iloc[ed])
                if abs(event.xdata - xs) < thresh:
                    menu = QtWidgets.QMenu(self)
                    menu.addAction('Delete start', lambda i=idx: self.delete_boundary(i, True))
                    menu.exec_(QtGui.QCursor.pos()); return
                if abs(event.xdata - xe) < thresh:
                    menu = QtWidgets.QMenu(self)
                    menu.addAction('Delete end', lambda i=idx: self.delete_boundary(i, False))
                    menu.exec_(QtGui.QCursor.pos()); return
            menu = QtWidgets.QMenu(self)
            menu.addAction('Add start', lambda: self.add_boundary_at(event, 'start'))
            menu.addAction('Add end', lambda: self.add_boundary_at(event, 'end'))
            menu.exec_(QtGui.QCursor.pos())

    def delete_boundary(self, idx, is_start):
        if 0 <= idx < len(self.segments):
            del self.segments[idx]
            self._draw_plot()

    def add_boundary_at(self, event, mode):
        dt = mdates.num2date(event.xdata)
        if hasattr(dt, 'tzinfo') and dt.tzinfo: dt = dt.replace(tzinfo=None)
        idx = np.abs(self.df['datetime'] - pd.Timestamp(dt)).idxmin()
        if mode == 'start':
            self.segments.append((idx, idx))
        else:
            if self.segments and self.segments[-1][0] == self.segments[-1][1]:
                st, _ = self.segments[-1]
                self.segments[-1] = (st, idx)
            else:
                self.segments.append((idx, idx))
        self._draw_plot()

    def update_plot(self):
        if not self.df.empty:
            self._draw_plot()

    def export_selection(self):
        if self.df.empty:
            QMessageBox.warning(self, 'No data', 'Load data.')
            return
        try:
            s = pd.to_datetime(self.start_edit.text()); e = pd.to_datetime(self.end_edit.text())
        except:
            QMessageBox.critical(self, 'Invalid dates', 'Invalid data format.')
            return
        vis = [(st, ed) for st, ed in self.segments if self.df.datetime.iloc[st] >= s and self.df.datetime.iloc[ed] <= e]
        if not vis:
            QMessageBox.information(self, 'No segments', 'No segments in selected time period.')
            return

        # --- limit na maximálně 6 vzorků ---
        if len(vis) > 6:
            QMessageBox.warning(
                self,
                'Too many segments',
                'You can export just 6 segments.'
            )
            return

        # select template and output path
        outp, _ = QFileDialog.getSaveFileName(self, 'Save file', '', 'Excel (*.xlsx)')
        if not outp:
            return
        self.output_path = outp

        # load workbook once
        wb = load_workbook(self.template_path)
        ws_orig = wb.active
        # write original intervals to active sheet
        cols = [2 + 4*i for i in range(len(vis))]
        start_row = 3
        for i, (st, ed) in enumerate(vis):
            seg = self.df.loc[st:ed, ['datetime', 'síla', 'deformace']].reset_index(drop=True)
            for r, row in enumerate(seg.itertuples(False), start=start_row):
                ws_orig.cell(row=r, column=cols[i], value=row.datetime.strftime('%d.%m.%Y'))
                ws_orig.cell(row=r, column=cols[i]+1, value=row.datetime.strftime('%H:%M:%S'))
                ws_orig.cell(row=r, column=cols[i]+2, value=row.síla)
                ws_orig.cell(row=r, column=cols[i]+3, value=row.deformace)


        # přejmenujeme list "Sample" podle názvu testu
        if 'Sample' in wb.sheetnames:
            sample_ws = wb['Sample']
            # Použijeme název testu pro přejmenování listu
            test_name = getattr(self, 'test_name', 'Unnamed Test')  # Pokud není název testu, použijeme výchozí název
            sample_ws.title = test_name

        # write resampled intervals as new sheets
        for i, (st, ed) in enumerate(vis):
            seg = self.df.loc[st:ed, ['datetime', 'síla', 'deformace']].reset_index(drop=True)
            if len(seg) < 2:
                continue
            dt_interp = pd.date_range(start=seg.datetime.iloc[0], end=seg.datetime.iloc[-1], periods=500)
            old_idx = np.linspace(0, 1, len(seg))
            new_idx = np.linspace(0, 1, 500)
            force_new = np.interp(new_idx, old_idx, seg['síla'])
            deform_new = np.interp(new_idx, old_idx, seg['deformace'])
            ws = wb.create_sheet(title=f'Interval_{i+1}')
            ws.cell(1, 1, 'Date'); ws.cell(1, 2, 'Time'); ws.cell(1, 3, 'Síla'); ws.cell(1, 4, 'Deformace')
            for r, (dt_val, f_val, d_val) in enumerate(zip(dt_interp, force_new, deform_new), start=2):
                ws.cell(r, 1, dt_val.strftime('%d.%m.%Y'))
                ws.cell(r, 2, dt_val.strftime('%H:%M:%S.%f')[:-3])
                ws.cell(r, 3, f_val)
                ws.cell(r, 4, d_val)
        print("▶ export: after resampled Interval sheets")

        if self.sample_params is not None:
            # odstraňte starý list, pokud existuje
            if 'SampleParameters' in wb.sheetnames:
                wb.remove(wb['SampleParameters'])
            sheet = wb.create_sheet(title='SampleParameters')
            # hlavička
            for j, col in enumerate(self.sample_params.columns, start=1):
                sheet.cell(row=1, column=j, value=col)
            # data
            for i, row in enumerate(self.sample_params.itertuples(index=False), start=2):
                for j, val in enumerate(row, start=1):
                    sheet.cell(row=i, column=j, value=val)

        if self.sample_params is not None and 'Sample' in wb.sheetnames:
            ws_sample = wb['Sample']
            # iterujeme řádky DataFrame, první se má zapsat do řádku 4
            for idx, row in enumerate(self.sample_params.itertuples(index=False), start=4):
                # Parametry by měly být ve sloupcích A, B, C, D, E (příp. F)
                ws_sample[f'CQ{idx}'] = row[0]   # A → CQ
                ws_sample[f'CR{idx}'] = row[1]   # B → CR
                ws_sample[f'CS{idx}'] = row[2]   # C → CS
                ws_sample[f'CT{idx}'] = row[3]   # D → CT
                ws_sample[f'CX{idx}'] = row[4]   # E → CX
                ws_sample[f'DK{idx}'] = row[5]   # F → DK
        print("▶ export: after SampleParameters")


        print("▶ summary_data blok: start")
        summary_data = []
        for name in wb.sheetnames:
            print(f"▶ summary_data: checking {name}")
            if not name.startswith('Interval_'):
                continue
            print(f"▶ summary_data: processing {name}")
            ws = wb[name]

            # robustní čtení
            all_rows = list(ws.values)
            if len(all_rows) < 2:
                print(f"▶ skip {name}: nemá žádná data")
                continue
            cols = all_rows[0]
            df_int = pd.DataFrame(all_rows[1:], columns=cols)

            # zpracujeme
            df_proc, used_deg, infl_time, infl_force, infl_deform, infl_force_fit, infl_deform_fit, *_ = process_interval_df(df_int)
            summary_data.append((name, used_deg, infl_time, infl_force, infl_deform, infl_force_fit, infl_deform_fit))

            # bezpečné mazání
            rows_to_remove = ws.max_row - 1
            if rows_to_remove > 0:
                ws.delete_rows(2, rows_to_remove)

            # zápis nového DF
            for r_idx, row in enumerate(dataframe_to_rows(df_proc, index=False, header=True), start=1):
                for c_idx, val in enumerate(row, start=1):
                    ws.cell(row=r_idx, column=c_idx, value=val)

        print("▶ summary_data blok: end")

        if 'Sample' in wb.sheetnames:
            sample_ws = wb['Sample']
            # najdi infl_force v summary_data
            inf_force = next((item[3] for item in summary_data if item[0] == 'Interval_1'), None)
            if inf_force is not None:
                sample_ws.cell(
                    row=53,
                    column=column_index_from_string('DG'),
                    value=inf_force
                )

        # — 2) Přepiš nebo vytvoř list 'Summary' se všemi inflexními body —
        if 'Summary' in wb.sheetnames:
            wb.remove(wb['Summary'])
        summary_ws = wb.create_sheet(title='Summary')
        headers = ['List', 'Used_Degree', 'Infl_Time', 'Infl_Force', 'Infl_Deform', 'Infl_Force_Fit', 'Infl_Deform_Fit']
        for col_idx, header in enumerate(headers, start=1):
            summary_ws.cell(row=1, column=col_idx, value=header)
        for row_idx, (sheet, deg, t, f, d, ff, df_f) in enumerate(summary_data, start=2):
            summary_ws.cell(row=row_idx, column=1, value=sheet)
            summary_ws.cell(row=row_idx, column=2, value=deg)
            summary_ws.cell(row=row_idx, column=3, value=t)
            summary_ws.cell(row=row_idx, column=4, value=f)
            summary_ws.cell(row=row_idx, column=5, value=d)
            summary_ws.cell(row=row_idx, column=6, value=ff)
            summary_ws.cell(row=row_idx, column=7, value=df_f)

        #  Přepis inflexních bodů do listu Sample jen pokud je zvolena auto detekce
        if self.radio_auto.isChecked() and 'Sample' in wb.sheetnames:
            ws_sample = wb['Sample']
            # 1) fitované body do řádků 38–43 (DG = síla_fit, DF = deform_fit)
            for idx, (_, _, _, _, _, infl_force_fit, infl_deform_fit) in enumerate(summary_data, start=2):
                row_excel = 36 + idx
                ws_sample.cell(row=row_excel,
                               column=column_index_from_string('DG'),
                               value=infl_force_fit)
                ws_sample.cell(row=row_excel,
                               column=column_index_from_string('DF'),
                               value=infl_deform_fit)

        print("▶ export: before save")
        # save workbook
        # remove old Interval_ sheets
        for name in list(wb.sheetnames):
            if name.startswith('Interval_') or name == 'SampleParameters' or name == 'Summary':
                wb.remove(wb[name])
        wb.save(self.output_path)
        print("▶ export: after save")
        QMessageBox.information(self, 'Done', f'Exported: {self.output_path}')

    def edit_parameters(self):
        if not self.segments:
            QMessageBox.warning(self, 'No segments', 'Nejdřív načtěte/namapujte intervaly.')
            return

        max_rows = 6
        total = len(self.segments)
        if total > max_rows:
            QMessageBox.information(
                self,
                'Limit parametrů',
                f'Máte {total} intervalů, v tabulce bude ale jen prvních {max_rows}.'
            )

        # vytvoříme DataFrame s indexy podle segmentů a 5 sloupců pro parametry
        n = min(total, max_rows)
        cols = ["Force", "Thickness", "Diameter", "Height", "Temperature", "Notch depth"]
        df = pd.DataFrame(np.nan, index=range(n), columns=cols)
        # pokud už jsme někdy parametry ukládali, předvyplníme
        if self.sample_params is not None and len(self.sample_params) == n:
            df[:] = self.sample_params.values

        # QDialog s QTableWidget
        dialog = QtWidgets.QDialog(self)
        dialog.setWindowTitle('Edit Sample Parameters')
        font = QtGui.QFont()
        font.setPointSize(9)
        dialog.setFont(font)
        dlg_layout = QtWidgets.QVBoxLayout(dialog)
        dialog.resize(700, 500)  # nebo dialog.setMinimumSize(800,500)

        # Přidáme textové pole pro název testu
        test_name_label = QtWidgets.QLabel("Test Name:")
        self.test_name_edit = QtWidgets.QLineEdit(dialog)  # Nové textové pole pro název testu
        dlg_layout.addWidget(test_name_label)
        dlg_layout.addWidget(self.test_name_edit)

        table = QtWidgets.QTableWidget(n, len(cols))
        table.setFont(font)  # ujistíme se, že tabulka také má 9pt font
        table.verticalHeader().setDefaultSectionSize(24)

        table.setHorizontalHeaderLabels(cols)
        # naplníme tabulku
        for i in range(n):
            for j, col in enumerate(cols):
                val = df.at[i, col]
                item = QtWidgets.QTableWidgetItem("" if pd.isna(val) else str(val))
                table.setItem(i, j, item)

        btn_import = QtWidgets.QPushButton("Import from Excel")

        def on_import():
            new_df = self.import_parameters_from_excel()
            if new_df is not None:
                for i in range(min(n, len(new_df))):
                    for j in range(min(len(cols), len(new_df.columns))):
                        val = new_df.iat[i, j]
                        item = QtWidgets.QTableWidgetItem("" if pd.isna(val) else str(val))
                        table.setItem(i, j, item)

        dlg_layout.addWidget(btn_import)
        btn_import.clicked.connect(on_import)

        dlg_layout.addWidget(table)

        # OK / Cancel tlačítka
        btns = QtWidgets.QDialogButtonBox(
            QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel)
        dlg_layout.addWidget(btns)
        btns.accepted.connect(dialog.accept)
        btns.rejected.connect(dialog.reject)

        if dialog.exec_() == QtWidgets.QDialog.Accepted:
            # načteme zpět DataFrame
            for i in range(n):
                for j, col in enumerate(cols):
                    item = table.item(i, j)
                    try:
                        df.at[i, col] = float(item.text())
                    except:
                        df.at[i, col] = item.text()
            self.sample_params = df
            # Získání názvu testu z textového pole
            test_name = self.test_name_edit.text()
            self.test_name = test_name  # Uložení názvu testu pro pozdější použití

    def import_parameters_from_excel(self):
        path, _ = QFileDialog.getOpenFileName(self, 'Load Excel File', '', 'Excel Files (*.xlsx *.xls)')
        if not path:
            return None

        try:
            df_all = pd.read_excel(path, sheet_name=None)
            sheet_names = list(df_all.keys())
            selected, ok = QtWidgets.QInputDialog.getItem(self, 'Select Sheet', 'Choose sheet to load:', sheet_names,
                                                          editable=False)
            if not ok or selected not in df_all:
                return None

            df_sheet = df_all[selected]

            # Získání řádku a sloupce, od kterých se bude číst
            start_row, ok1 = QtWidgets.QInputDialog.getInt(self, 'Start Row', 'Enter starting row (e.g., 2):', value=2,
                                                           min=1)
            if not ok1:
                return None
            start_col, ok2 = QtWidgets.QInputDialog.getInt(self, 'Start Column', 'Enter starting column (e.g., 2 = B):',
                                                           value=2, min=1)
            if not ok2:
                return None

            # Načíst data – max 6 řádků a 6 sloupců
            max_rows, max_cols = 6, 6
            df_sheet = df_sheet.iloc[start_row - 1:start_row - 1 + max_rows, start_col - 1:start_col - 1 + max_cols]

            # Nastavení hlaviček
            cols = ["Force", "Thickness", "Diameter", "Height", "Temperature", "Notch depth"]
            df_sheet.columns = cols[:df_sheet.shape[1]]

            return df_sheet.reset_index(drop=True)

        except Exception as e:
            QMessageBox.critical(self, 'Import Error', str(e))
            return None


class DraggableVLine:
    def __init__(self, line, seg_idx, is_start, app):
        self.line = line; self.seg_idx = seg_idx; self.is_start = is_start; self.app = app
        self.press = None; self.orig_xlim = None
        self.cid_press = app.canvas.mpl_connect('button_press_event', self.on_press)
        self.cid_motion = app.canvas.mpl_connect('motion_notify_event', self.on_motion)
        self.cid_release = app.canvas.mpl_connect('button_release_event', self.on_release)

    def on_press(self, event):
        contains, _ = self.line.contains(event)
        if not contains:
            return
        x0 = self.line.get_xdata()[0]
        xp = event.x  # pixelová souřadnice
        self.press = (x0, xp)
        self.orig_xlim = self.app.canvas.ax.get_xlim()

    def on_motion(self, event):
        if self.press is None or event.x is None:
            return
        ax = self.app.canvas.ax
        x0, xp = self.press
        width_pixels = ax.get_window_extent().width
        dx_pix = event.x - xp
        span = self.orig_xlim[1] - self.orig_xlim[0]
        dx_data = dx_pix / width_pixels * span
        nx = x0 + dx_data
        self.line.set_xdata([nx, nx])
        if self.app.chk_zoom.isChecked():
            zs = span * 0.2
            ax.set_xlim(nx - zs / 2, nx + zs / 2)
        self.app.canvas.draw()

    def on_release(self, event):
        if self.press is None: return
        nx = self.line.get_xdata()[0]
        dt = mdates.num2date(nx)
        if hasattr(dt, 'tzinfo') and dt.tzinfo: dt=dt.replace(tzinfo=None)
        idx = np.abs(self.app.df['datetime'] - pd.Timestamp(dt)).idxmin()
        s, e = self.app.segments[self.seg_idx]
        if self.is_start: self.app.segments[self.seg_idx] = (idx, max(idx+1, e))
        else: self.app.segments[self.seg_idx] = (min(s, idx-1), idx)
        self.app.canvas.ax.set_xlim(self.orig_xlim)
        self.press=None; self.app._draw_plot()

if __name__ == '__main__':
    app = QtWidgets.QApplication(sys.argv)
    win = SCBApp()
    win.show()
    sys.exit(app.exec_())
